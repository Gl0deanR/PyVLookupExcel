from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook

column_cell_1 = ""
column_cell_2 = ""
column_stock_1 = ""
column_stock_2 = ""
row_file_1 = ""
row_file_2 = ""
path1 = ""
path2 = ""


def get_file_path1():
    # Open and return file path
    global path1
    file_path1 = filedialog.askopenfilename(title="Choose an excel file", filetypes=[("Excel files", ".xlsx .xls")])
    path1 = file_path1


def get_file_path2():
    # Open and return file path
    global path2
    file_path2 = filedialog.askopenfilename(title="Choose an excel file", filetypes=[("Excel files", ".xlsx .xls")])
    path2 = file_path2


def get_columns():
    global column_cell_1, column_cell_2, column_stock_1, column_stock_2, row_file_1, row_file_2
    column_cell_1 = column_1.get()
    column_cell_2 = column_2.get()
    column_stock_1 = e_column_1.get()
    column_stock_2 = e_column_2.get()
    row_file_1 = e_rand_1.get()
    row_file_2 = e_rand_2.get()


window = Tk()
window.title('VLookup by Gl0deanR')
window.geometry("400x400")

# get the column numbers that have the unique values in both files, for ex if they have a unique product code you would choose the "Product Code" columns
Label(window, text="Number of unique column in file 1:\t").grid(row=2)
Label(window, text="Number of unique column in file 2:\t").grid(row=3)

# get the row numbers of the values that need to be searched and written, for ex if you need the stock values you would choose the "Stock" columns
Label(window, text="Number of the column to be searched in file 1:\t").grid(row=4)
Label(window, text="Number of the column to be written in file 2:\t").grid(row=5)

# get the row numbers of each files where the data starts; this is just for optimizing reasons
Label(window, text="First row of file 1:\t").grid(row=6)
Label(window, text="First row of file 2:\t").grid(row=7)

column_1 = Entry(window)
column_1.grid(row=2, column=1)
column_2 = Entry(window)
column_2.grid(row=3, column=1)
e_column_1 = Entry(window)
e_column_1.grid(row=4, column=1)
e_column_2 = Entry(window)
e_column_2.grid(row=5, column=1)
e_rand_1 = Entry(window)
e_rand_1.grid(row=6, column=1)
e_rand_2 = Entry(window)
e_rand_2.grid(row=7, column=1)

# you press the SEND button then close the GUI window
button_get_columns = Button(window, text="SEND", command=get_columns)
button_get_columns.grid(row=9, column=1)

button_get_path1 = Button(window, text="Choose file 1", command=get_file_path1)
button_get_path1.grid(row=0)
button_get_path2 = Button(window, text="Choose file 2", command=get_file_path2)
button_get_path2.grid(row=1)

window.mainloop()

column_cell_1 = int(column_cell_1)
column_cell_2 = int(column_cell_2)
column_stock_1 = int(column_stock_1)
column_stock_2 = int(column_stock_2)
row_file_1 = int(row_file_1)
row_file_2 = int(row_file_2)

wb_obj1 = load_workbook(path1, data_only=True)
sheet_obj1 = wb_obj1.active
wb_obj2 = load_workbook(path2, data_only=True)
sheet_obj2 = wb_obj2.active

for i in range(row_file_1, sheet_obj1.max_row + 1):
    cell_obj1 = sheet_obj1.cell(row=i, column=column_cell_1)
    if cell_obj1.value is None or cell_obj1.value == 0:
        continue
    if sheet_obj1.row_dimensions[i].hidden:
        continue
    for ii in range(row_file_2, sheet_obj2.max_row + 1):
        cell_obj2 = sheet_obj2.cell(row=ii, column=column_cell_2)
        if cell_obj1.value == cell_obj2.value:
            cell_obj_final = sheet_obj2.cell(row=ii, column=column_stock_2)
            cell_obj1_de_luat = sheet_obj1.cell(row=i, column=column_stock_1)
            cell_obj_final.value = cell_obj1_de_luat.value


wb_obj2.save(path2)

window_success = Tk()
window_success.title('VLookup by Gl0deanR')
window_success.geometry("400x400")

Label(window_success, text="PROCESS FINISHED WITH SUCCESS!").grid(row=2)

window_success.mainloop()
